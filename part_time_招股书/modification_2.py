import openpyxl
import os
import re

def is_sheet_empty(worksheet):
    # 检查是否是空表
    if worksheet.dimensions == 'A1:A1' and worksheet['A1'].value is None:
        return True
    return False

def choose_title(title_candidate:list):
    filtered = []
    for s in title_candidate:
        if  re.search(r'\(|\d+', s) and '单位' not in s:
            filtered.append(s)
    if filtered:
        if len(filtered) >= 2:
            if len(filtered[-1]) < len(filtered[-2]):
                result = filtered[-1]
            else:
                result = filtered[-2]
        else: result = filtered[-1]
        return result
    else:
        #raise ValueError("No matching string found!")
        return 'UNK'
        



def modification_2(file,output):
    os.makedirs(os.path.dirname(output), exist_ok=True)
    workbook = openpyxl.load_workbook(file)
    if not is_sheet_empty(workbook['主营业务收入，按产品、地域等划分']):
        old_sheet = workbook['主营业务收入，按产品、地域等划分']
    else :
        print('Empty sheet!')
        workbook.save(output)
        debug = (os.path.basename(file), 'Empty sheet!')
        return debug
    
    # newbook = openpyxl.Workbook()
    # sheet = newbook.active
    # sheet.title = '主营业务收入，按产品、地域等划分'
    sheet = old_sheet
    sheet.title = '主营业务收入，按产品、地域等划分new'
    for row in old_sheet.iter_rows():
        for cell in row:
            sheet[cell.coordinate].value = cell.value
    sheet.insert_cols(1)

    division_titles = [0]
    buffer = []
    prev_division_title = ''
    for r in range(1,sheet.max_row+1): 
        c1 = sheet.cell(row = r,column=2).value
        c2 = sheet.cell(row = r,column=3).value
        c3 = sheet.cell(row = r,column=4).value
        if c1 and not c2 :
            buffer.append(c1)
        try:
            if (c1 and c2) or (c2 and c3):
                if buffer:
                    division_titles = buffer
                d_title = choose_title(division_titles)
                if len(d_title) > 20 and len(division_titles) < 4:
                    division_title = prev_division_title
                else:
                    division_title = d_title
                    prev_division_title = division_title
                
                division_title = f'table_{division_title}'
                sheet.cell(row = r,column= 1).value = division_title
                buffer = []
        except Exception as e:
            print(f'{os.path.basename(file)} : {e}')
            debug = (os.path.basename(file),str(e))
            workbook.save(output)
            return debug 
    try:
        workbook.save(output)
    except Exception as e:
        print(f'{os.path.basename(file)} : {e}')
        debug = (os.path.basename(file),str(e))
        return debug 

if __name__ == '__main__':
    pass