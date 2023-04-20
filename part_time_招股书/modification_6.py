import openpyxl
import os
import re

def is_sheet_empty(worksheet):
    # 检查是否是空表
    if worksheet.dimensions == 'A1:A1' and worksheet['A1'].value is None:
        return True
    return False

def choose_title(title_candidate:list):
    title_candidate = [s for s in title_candidate if len(s) > 5] 
    if len(title_candidate) >= 2:
        candidate = title_candidate[-2:]
        candidate_ = [s for s in candidate if '如下' in s or '报告期' in s]
        return candidate_[-1] if candidate_ else candidate[-1]
    elif len(title_candidate) == 1:
        return title_candidate[0]
    else:
        #raise ValueError("No matching string found!")
        return 'UNK'
        
def choose_title(title_candidate:list):
    title_candidate = [s for s in title_candidate if len(s) > 8]
    _ = [s for s in title_candidate if '如下' in s]
    patterns = [r"\(\d+\)", r"\d、", r"\(.*\d.*\)",r"\([一二三四五六七八九十]+\)", r"\d+、",r'\d 、',r'[一二三四五六七八九十]、',r"\（[一二三四五六七八九十]+\）"]
    matching_elements = []
    for element in title_candidate:
        for pattern in patterns:
            if re.search(pattern, element[:10]):
                matching_elements.append(element)
                break
    if matching_elements:
        return matching_elements[-1]
    elif _:
        return _[-1]    
    else:
        return 'UNK'        



def modification_6(file,output):
    os.makedirs(os.path.dirname(output), exist_ok=True)
    workbook = openpyxl.load_workbook(file)
    if  '前十大股东' not in workbook.sheetnames:
        workbook.save(output)
        print(f'{os.path.basename(file)} : Missing sheet!')
        debug = (os.path.basename(file), 'Missing sheet!')
        return debug
    elif not is_sheet_empty(workbook['前十大股东']):
        old_sheet = workbook['前十大股东']
    else :
        print('Empty sheet!')
        workbook.save(output)
        debug = (os.path.basename(file), 'Empty sheet!')
        return debug

    # newbook = openpyxl.Workbook()
    # sheet = newbook.active
    sheet = old_sheet
    sheet.title = '前十大股东new'
    for row in old_sheet.iter_rows():
        for cell in row:
            sheet[cell.coordinate].value = cell.value
    sheet.insert_cols(1)

    division_titles = []
    prev_division_title = ''
    years = []
    buffer = []
    quit_maker = 0
    for r in range(1,sheet.max_row+1):
        c1 = sheet.cell(row = r,column=2).value
        c2 = sheet.cell(row = r,column=3).value
        c3 = sheet.cell(row = r,column=4).value
        c4 = sheet.cell(row = r,column=5).value
        if c1 and not c2:
            buffer.append(c1)
            if quit_maker == 1:
                workbook.save(output)
                return
        try:
            if (c1 and c2) or (c2 and c3) or (c3 and c4):
                quit_maker = 1
                if buffer:
                    division_titles = buffer
                division_title = choose_title(division_titles)
                #print(division_title)
                if division_title == 'UNK':
                    d_title = prev_division_title
                else:
                    d_title = division_title
                    prev_division_title = division_title
                division_title = f'table_{d_title}'
                sheet.cell(row = r,column= 1).value = division_title
                buffer = []
        except Exception as e:
            workbook.save(output)
            print(f'{os.path.basename(file)} : {e}')
            debug = (os.path.basename(file),str(e))
            return debug 

    try:
        workbook.save(output)
    except Exception as e:
        print(f'{os.path.basename(file)} : {e}')
        debug = (os.path.basename(file),str(e))
        return debug


if __name__ == '__main__':
    modification_6("x.xlsx","ST55.SH.xlsx")