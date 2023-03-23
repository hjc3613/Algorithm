import openpyxl
import os


def is_sheet_empty(worksheet):
    # 检查是否是空表
    if worksheet.dimensions == 'A1:A1' and worksheet['A1'].value is None:
        return True
    return False


def modification_1(file,output):
    os.makedirs(os.path.dirname(output), exist_ok=True)
    workbook = openpyxl.load_workbook(file)
    if '员工专业学历' in workbook.get_sheet_names() and not is_sheet_empty(workbook['员工专业学历']):
        old_sheet = workbook['员工专业学历']
    else :
        debug = (os.path.basename(file), 'Empty sheet!')
        workbook.save(output)
        return debug
    # newbook = openpyxl.Workbook()
    # sheet = newbook.active
    sheet = workbook.create_sheet('员工专业学历new')
    # sheet.title = '员工专业学历'
    for row in old_sheet.iter_rows():
        for cell in row:
            sheet[cell.coordinate].value = cell.value
    sheet.insert_cols(1)
    #关键字
    edu = ['本科','硕士']
    skils = ['人员','销售']
    age = ['岁']
    subdivision = ''
    prev_marker = 1
    conclude = ['总计','合计']
    for r in range(1,sheet.max_row+1): 
        c1 = sheet.cell(row = r,column=2).value or ''
        c2 = sheet.cell(row = r,column=3).value
        c3 = sheet.cell(row = r,column=4).value
        # try:
        #通过关键字决定标注内容
        if any([x in c1 for x in skils]):
            subdivision = "table_专业结构"
        elif any([x in c1 for x in edu]):
            subdivision = "table_教育程度"
        elif any([x in c1 for x in age]):
            subdivision = "table_年龄结构"
        #碰到‘合计’关键字后开始标准前面读过的行
        if any([x in c1 for x in conclude]):
            for prev_row in range(prev_marker,r):
                sub_c1 = sheet.cell(row = prev_row+1,column= 2).value
                sub_c2 = sheet.cell(row = prev_row+1,column= 3).value
                sub_c3 = sheet.cell(row = prev_row+1,column= 4).value
                #标注条件为第1，2列不为空
                if sub_c1 and sub_c2:
                    sheet.cell(row = prev_row+1,column= 1).value = subdivision
            prev_marker = r
        # except Exception as e:
        #     #print(f'{os.path.basename(file)} : {e}')
        #     debug = (os.path.basename(file),str(e))
        #     workbook.save(output)
        #     return debug
    # newbook.save(output)
    workbook.save(output)
    



if __name__ == '__main__':
    pass