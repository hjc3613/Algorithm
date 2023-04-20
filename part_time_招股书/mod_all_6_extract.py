import openpyxl
import os
import re
import json

directory = '最新文件抽取结果-fix5'
output_dir = '最新文件抽取结果-fix6'
os.makedirs(output_dir, exist_ok=True)

# Create a new workbook and select the worksheet
converted_wb = openpyxl.Workbook()
converted_ws = converted_wb.active
converted_ws.insert_cols(1, 4)
column_names = ["stock_id", "table", "subject", "value"]
for i, name in enumerate(column_names):
    converted_ws.cell(row=1, column=1 + i).value = name


def extract_single(original_ws, converted_ws, stock_id):
    rows_so_far = converted_ws.max_row
    sub_table_buffer = ""
    row_count = 0
    # Iterate through the original table rows and convert them
    for r in range(1, original_ws.max_row + 1):
        coloumn = [cell.value for cell in original_ws[r]]
        if not coloumn:
            break
        if coloumn[0]:

            sub_table = coloumn[0]

            if sub_table != sub_table_buffer:
                coloumn[1:] = [c for c in coloumn[1:] if c != None]
                subjects = [c for c in coloumn[1:] if c != "序号" and c != None]
                no_ids_len = len(coloumn[1:]) - len(subjects)
                sub_table_buffer = sub_table
            else:
                try:
                    coloumn_content = [c for c in coloumn[1:] if c != "序号" and c != None]
                    if coloumn_content == subjects:
                        continue
                    elif "合计" in coloumn_content or "总计" in coloumn_content:
                        coloumn_content = [c for c in coloumn_content if c not in ["合计", "总计"]]
                        for row, value in enumerate(coloumn_content, start=1):
                            converted_ws.cell(row=row + row_count + rows_so_far, column=1).value = stock_id
                            converted_ws.cell(row=row + row_count + rows_so_far, column=2).value = sub_table
                            converted_ws.cell(row=row + row_count + rows_so_far, column=3).value = '合计/合计(占比)'
                            converted_ws.cell(row=row + row_count + rows_so_far, column=4).value = value

                    else:
                        coloumn_content = coloumn_content[no_ids_len:]
                        for row, value in enumerate(coloumn_content, start=1):
                            converted_ws.cell(row=row + row_count + rows_so_far, column=1).value = stock_id
                            converted_ws.cell(row=row + row_count + rows_so_far, column=2).value = sub_table
                            converted_ws.cell(row=row + row_count + rows_so_far, column=3).value = subjects[row - 1]
                            converted_ws.cell(row=row + row_count + rows_so_far, column=4).value = value
                    row_count = row_count + len(coloumn_content)
                except Exception as e:
                    print(f"Error in {stock_id} {sub_table} {coloumn} as {e}")
                    debug = f"{stock_id} : {str(e)})"
                    return debug


file_path = []
ignore = ['__pycache__', 'no_dup', '主营业务收入，按产品、地域等划分',
          '员工专业学历', '第六节业务与技术', '前五大客户销售', '经营成果分析', '毛利率']
for root, dirs, files in os.walk(directory):
    for ig in ignore:
        if ig in dirs:
            dirs.remove(ig)
    for filename in files:
        # if not any([x in filename for x in ['py','txt','json','zip']]):
        #     file_path.append(os.path.join(root, filename))
        if filename.endswith('.xlsx'):
            file_path.append(os.path.join(root, filename))
    file_path = [f for f in file_path if '前十大股东' in f]

number_pattern = re.compile(r'\d+')
debugs = {}
# file_path = ['豪森股份(688529.SH).xlsx']
for i, f in enumerate(file_path):
    print(f"Processing item {i + 1}/{len(file_path)}: {f}", end='\r')
    dir = os.path.dirname(f)
    file_name = os.path.basename(f)
    stock_id = number_pattern.search(file_name).group()
    original_wb = openpyxl.load_workbook(f)
    original_ws = original_wb.active
    debug = extract_single(original_ws, converted_ws, stock_id)
    if debug:
        debugs[stock_id] = debug
    original_wb.close()
converted_wb.save('converted_仅前十股东.xlsx')
if debugs:
    with open('debug_仅前十股东.json', 'w') as f:
        json.dump(debugs, f, indent=4, ensure_ascii=False)

