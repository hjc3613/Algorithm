import openpyxl
import os
import re

def is_sheet_empty(worksheet):
    # 检查是否是空表
    if worksheet.dimensions == 'A1:A1' and worksheet['A1'].value is None:
        return True
    return False

def choose_title(title_candidate:list):
    #filtered = []
    #for s in title_candidate:
        #if  "客户" in s  and len(s) < 30:
            #filtered.append(s)
    #if filtered:
    title_candidate = [s for s in title_candidate if "单位"  not in s and '招股' not in s and len(s) > 10]
    if title_candidate:
        #return filtered[-1]
        return min(title_candidate, key=len)
    else:
        #raise ValueError("No matching string found!")
        return 'UNK'
        



def modification_3(file,output):
    os.makedirs(os.path.dirname(output), exist_ok=True)
    workbook = openpyxl.load_workbook(file)
    if  '前五大客户销售' not in workbook.sheetnames:
        print(f'{os.path.basename(file)} : Missing sheet!')
        debug = (os.path.basename(file), 'Missing sheet!')
        workbook.save(output)
        return debug
    elif not is_sheet_empty(workbook['前五大客户销售']):
        old_sheet = workbook['前五大客户销售']
    else :
        print('Empty sheet!')
        workbook.save(output)
        debug = (os.path.basename(file), 'Empty sheet!')
        return debug
    
    # newbook = openpyxl.Workbook()
    # sheet = newbook.active
    sheet = old_sheet
    sheet.title = '前五大客户销售new'
    for row in old_sheet.iter_rows():
        for cell in row:
            sheet[cell.coordinate].value = cell.value
    sheet.insert_cols(1)

    division_titles = []
    prev_division_title = ''
    years = []
    buffer = []
    for r in range(1,sheet.max_row+1): 
        c1 = sheet.cell(row = r,column=2).value
        c2 = sheet.cell(row = r,column=3).value
        c3 = sheet.cell(row = r,column=4).value
        if (c1 and '年' not in c1 ) and not c2 :
            buffer.append(c1)
        try:
            if (c1 and c2) or (c2 and c3) or (c1 and '年' in c1 and len(c1) < 20):
                if buffer:
                    division_titles = buffer
                division_title = choose_title(division_titles)
                #print(division_title)
                if division_title == 'UNK':
                    d_title = prev_division_title
                else:
                    d_title = division_title
                    prev_division_title = division_title
                division_title = f'table_{d_title}'
                sheet.cell(row = r,column= 1).value = division_title
                buffer = []
        except Exception as e:
            print(f'{os.path.basename(file)} : {e}')
            workbook.save(output)
            debug = (os.path.basename(file),str(e))
            return debug 
    try:
        workbook.save(output)
    except Exception as e:
        workbook.save(output)
        print(f'{os.path.basename(file)} : {e}')
        debug = (os.path.basename(file),str(e))
        return debug 

if __name__ == '__main__':
    modification_3('x.xlsx','1.xlsx')